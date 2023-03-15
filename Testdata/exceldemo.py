import openpyxl

book = openpyxl.load_workbook("C:\\Users\\pavij\\Documents\\pythondemo.xlsx")
sheet = book.active
Dict = {}
sheet.cell(row=1, column=2)
# print(cell.value)              #prints the value in the exel sheet i.e,read from excel sheet

sheet.cell(row=2, column=2).value = "pavithra"  # writes the value to excel sheet i.e, write to the excel sheet
print(sheet.cell(row=2, column=2).value)

print(sheet.max_row)            #total number of rows

print(sheet.max_column)               #total number of rows

print(sheet['A1'].value)       #prints the value in row3

#for i in range(1,sheet.max_row+1):
 #   for j in range(1,sheet.max_column+1):
  #      print(sheet.cell(row=i,column=j).value)   #print all the values in excelsheet

#for i in range(1,sheet.max_row+1):
 #   if sheet.cell(row=i,column=1).value=="test2":
  #      for j in range(1,sheet.max_column+1):
   #         print(sheet.cell(row=i,column=j).value)  #print all the data of test2

for i in range(1,sheet.max_row+1):
    if sheet.cell(row=i,column=1).value=="Test1":
        print(sheet.cell(row=i, column=1).value)
        for j in range(2,sheet.max_column+1):
            Dict[sheet.cell(row=1,column=j).value]=sheet.cell(row=i,column=j).value
print(Dict)

